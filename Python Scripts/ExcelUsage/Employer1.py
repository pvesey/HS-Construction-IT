from openpyxl import Workbook
import openpyxl

filename = "checklist_1_Vesey.xlsx"

workbook = openpyxl.load_workbook(filename)

sheet = workbook.active

cell = sheet.cell(row=6, column=2)

print(cell.value)

nullcell = sheet.cell(row=6, column=5)

if (nullcell.value is None):
    print("Empty")


sheet["D6"] = "Yes"
sheet["D7"] = "Yes"

workbook.save(filename)