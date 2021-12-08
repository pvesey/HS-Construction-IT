#Assignment1.py
from openpyxl import Workbook
import openpyxl
from shutil import copyfile

template = "rtw_form_sep21.xlsx"

users = ["Alice Adams", "Bob Baker", "Carol Clark", "Dan Davis", "Erin Evans", "Faythe Frank", "Grace Ghosh",
        "Heidi Hills", "Ivan Irwin", "John Jones", "Kevin Klein", "Lisa Lopez"]


# Approach 1 -- open the template file, make edits and save as new file name
def version1(template, users):
    for user in users:
        workbook = openpyxl.load_workbook(template)
        workbook.save(user + ".xlsx")
        filename = user + ".xlsx"
        userWorkbook = openpyxl.load_workbook(filename)
        sheet = userWorkbook.active
        sheet["A4"] = "Employee Name:  " + user
        sheet["A5"] = "Company Name: TUS Moylish Campus"
        sheet["C4"] = "Manager Name: Paul Vesey"
        userWorkbook.save(filename)


# Approach 2 -- Copy the file, then edit
def version2(template, users):
    for user in users:
        filename = user + ".xlsx"
        copyfile("rtw_form_sep21.xlsx", filename)
        userWorkbook = openpyxl.load_workbook(filename)
        sheet = userWorkbook.active
        sheet["A4"] = "Employee Name:  " + user
        sheet["A5"] = "Company Name: TUS Moylish Campus"
        sheet["C4"] = "Manager Name: Paul Vesey"
        userWorkbook.save(filename)


# Approach 3 -- Copy the files first and then edit each in turn.  Very similar to batch file approach
def version3(template, users):
    for user in users:
        filename = user + ".xlsx"
        copyfile("rtw_form_sep21.xlsx", filename)

    for user in users:
        filename = user + ".xlsx"
        copyfile("rtw_form_sep21.xlsx", filename)
        userWorkbook = openpyxl.load_workbook(filename)
        sheet = userWorkbook.active
        sheet["A4"] = "Employee Name:  " + user
        sheet["A5"] = "Company Name: TUS Moylish Campus"
        sheet["C4"] = "Manager Name: Paul Vesey"
        userWorkbook.save(filename)

    return "Version 3 Complete"




#version1(template, users)
#version2(template, users)
print(version3(template, users))
