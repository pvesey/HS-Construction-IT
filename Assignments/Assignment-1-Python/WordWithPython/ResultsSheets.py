import os, sys  # Standard Python Libraries
from docxtpl import DocxTemplate  # pip install docxtpl
import win32com.client as win32  # pip install pywin32
from openpyxl import Workbook
import openpyxl

wb = Workbook()

path = "StudentNames.xlsx"
 
wb = openpyxl.load_workbook(path)
 
xl_sheet = wb.active
max_row = xl_sheet.max_row
 
# Loop will print all values
# of first column


# -- Documentation:
# python-docx-template: https://docxtpl.readthedocs.io/en/latest/

# Change path to current working directory
os.chdir(sys.path[0])



def convert_to_pdf(doc):
    """Convert given word document to pdf"""
    word = win32.DispatchEx("Word.Application")
    new_name = doc.replace(".docx", r".pdf")
    worddoc = word.Documents.Open(doc)
    worddoc.SaveAs(new_name, FileFormat=17)
    worddoc.Close()
    return None


def main():
    wb = Workbook()

    path = "StudentNames.xlsx"
     
    wb = openpyxl.load_workbook(path)
     
    xl_sheet = wb.active
    max_row = xl_sheet.max_row

    dictionary = {}

    for row in xl_sheet.rows:
        doc = DocxTemplate("PythonForm.docx")
        output_name = f'test_{row[0].value}.docx'
        dictionary.update({"StudentName" : row[0].value})
        dictionary.update({"KNumber" : row[1].value})
        dictionary.update({"Comments" : row[2].value})
        doc.render(dictionary)
        doc.save(output_name)  
        print(output_name)
        dictionary.clear()

if __name__ == "__main__":
    main()