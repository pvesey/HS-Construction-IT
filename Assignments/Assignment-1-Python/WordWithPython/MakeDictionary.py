from openpyxl import Workbook
import openpyxl

wb = Workbook()

path = "StudentNames.xlsx"
 
wb = openpyxl.load_workbook(path)
 
xl_sheet = wb.active
max_row = xl_sheet.max_row
 
# Loop will print all values
# of first column

for row in xl_sheet.rows:
    dictionary = {}
    print(row[0].value)
    print(row[1].value)
    print(row[2].value)
    dictionary = {
        'StudentName' : row[0].value,
        'Knumber': row[1].value,
        'Comments' : row[2].value
    }
    print(dictionary)