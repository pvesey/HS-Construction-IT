from openpyxl import load_workbook
import datetime
employees_file = "EmployeeData.xlsx"

wb = load_workbook(filename = employees_file)

ws = wb['Sheet1']

my_list = []
my_row = []

for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        if (type(cell.value) is datetime.datetime):
            cell.value = cell.value.date()
        my_row.append(cell.value)
    my_list.append(my_row[:]) # the [:] is needed to ensure overwrite does not happen
    my_row.clear()


print(my_list)
print(len(my_list))

## NOW WE NEED TO CREATE THE DICTIONARY FROM THE DATA THAT WE HAVE IN THE LIST.

employee_dictionary = {}

for entry in my_list:
    temp_dict = {entry[0] : (entry[1], entry[2], entry[3])}
    employee_dictionary.update(temp_dict)

print(employee_dictionary)
print(employee_dictionary['EMP001'])


# NOW WE NEED TO READ IN THE SCAFF CHECKS FILE

checks_file = "ScaffChecks.xlsx"

wb2 = load_workbook(filename = checks_file)

ws = wb2['Sheet1']

my_list.clear()
my_row.clear()

for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        if (type(cell.value) is datetime.datetime):
            cell.value = cell.value.date()
        my_row.append(cell.value)
    my_list.append(my_row[:]) # the [:] is needed to ensure overwrite does not happen
    my_row.clear()

print(my_list)


count_of_pass = 0

for entry in my_list:
    if entry[4]:
        count_of_pass += 1

print('Count of Pass: ', count_of_pass)
